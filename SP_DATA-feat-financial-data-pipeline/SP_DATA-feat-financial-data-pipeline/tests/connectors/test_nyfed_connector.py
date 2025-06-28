import pytest
import pandas as pd
from datetime import date, datetime, timezone
import os
from unittest.mock import patch, MagicMock
from io import BytesIO
import openpyxl # For creating mock Excel data

from src.connectors.nyfed_connector import NYFedConnector

@pytest.fixture
def nyfed_config_minimal():
    """提供一個最小化的 NYFedConnector 設定，不含下載和解析配方。"""
    return {
        "requests_per_minute": 30
    }

@pytest.fixture
def nyfed_config_with_recipes():
    """提供一個帶有下載和解析配方的 NYFedConnector 設定。"""
    return {
        "requests_per_minute": 30,
        "download_configs": [
            {
                "name": "primary_dealer_stats_test",
                "url_template": "https://www.example.com/markets/prideal/prideal{YYYY}.xlsx",
                "file_pattern_on_page": "prideal{YYYY}.xlsx",
                "parser_recipe_name": "test_dealer_recipe",
                "metric_name_override": "NYFED/TEST_DEALER_METRIC"
            },
            { # 一個 HTML 頁面，需要從中找到連結的例子
                "name": "some_html_page_with_excel",
                "url_template": "https://www.example.com/data/some_page.html",
                "file_pattern_on_page": "specific_data_file_2023.xlsx", # 假設當年是2023
                "parser_recipe_name": "another_recipe_for_excel",
                "metric_name_override": "NYFED/HTML_SOURCED_METRIC"
            }
        ],
        "parser_recipes": {
            "test_dealer_recipe": {
                "header_row": 2, # 假設 Excel 標頭在第2行 (1-based)
                "date_column": "Report Date",
                "columns_to_sum": ["Metric A", "Metric B"],
                "data_unit_multiplier": 1,
                "sheet_name": "Sheet1"
            },
            "another_recipe_for_excel": {
                "header_row": 1,
                "date_column": "Date",
                "value_column": "Value", # 直接取值，不加總
                "sheet_name": 0 # 第一個 sheet
            }
        },
        "requests_config": { "max_retries": 1, "base_backoff_seconds": 0.1, "download_timeout": 10 }
    }

@pytest.fixture
def nyfed_connector(nyfed_config_with_recipes):
    """實例化一個 NYFedConnector。"""
    return NYFedConnector(api_config=nyfed_config_with_recipes)

# --- 輔助函數，用於創建模擬的 Excel BytesIO 物件 ---
def create_mock_excel_bytes(sheet_data_map: Dict[str, List[List[Any]]]) -> BytesIO:
    """
    創建一個包含指定數據的模擬 Excel 檔案的 BytesIO 物件。
    sheet_data_map: {"SheetName": [[row1_col1, row1_col2], [row2_col1, row2_col2], ...]}
    """
    workbook = openpyxl.Workbook()
    # 移除預設創建的 "Sheet"
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1 and not sheet_data_map.get("Sheet"):
         del workbook["Sheet"]

    first_sheet = True
    for sheet_name, data_rows in sheet_data_map.items():
        if first_sheet and sheet_name not in workbook.sheetnames:
            # openpyxl 預設會創建一個名為 'Sheet' 的工作表。
            # 如果我們的第一個 sheet 名稱不同，且 'Sheet' 仍存在，則重命名它。
            # 或者，如果我們只提供一個 sheet，且名稱不是 'Sheet'，則創建新 sheet 並刪除預設的。
            if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
                default_sheet = workbook["Sheet"]
                if not data_rows and sheet_name != "Sheet": # If default sheet is empty and we need a named one
                    workbook.create_sheet(title=sheet_name)
                    del workbook["Sheet"]
                elif data_rows: # We have data for a specific sheet name
                     if sheet_name != "Sheet":
                        workbook.create_sheet(title=sheet_name)
                        if "Sheet" in workbook.sheetnames and not workbook["Sheet"].calculate_dimension() == 'A1:A1': # if default sheet is not empty
                             pass # Keep default sheet if it has content or if we only want to add
                        elif "Sheet" in workbook.sheetnames:
                             del workbook["Sheet"] # Delete empty default sheet
                current_sheet = workbook[sheet_name]
            elif sheet_name not in workbook.sheetnames:
                 current_sheet = workbook.create_sheet(title=sheet_name)
            else: # sheet_name already exists
                 current_sheet = workbook[sheet_name]
            first_sheet = False
        elif sheet_name not in workbook.sheetnames:
            current_sheet = workbook.create_sheet(title=sheet_name)
        else:
            current_sheet = workbook[sheet_name]

        for row in data_rows:
            current_sheet.append(row)

    excel_bytes = BytesIO()
    workbook.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

# --- 測試 _download_file_content ---
@patch('requests.Session.get')
def test_download_file_direct_excel_success(mock_get, nyfed_connector):
    """測試直接下載 Excel 檔案成功。"""
    mock_excel_content = create_mock_excel_bytes({"Sheet1": [["Header"], [1]]})
    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.headers = {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
    mock_response.content = mock_excel_content.getvalue()
    mock_get.return_value = mock_response

    url = "https://www.example.com/direct_file.xlsx"
    result_bytesio = nyfed_connector._download_file_content(url)

    assert result_bytesio is not None
    assert result_bytesio.getvalue() == mock_excel_content.getvalue()
    mock_get.assert_called_once_with(url, timeout=nyfed_connector.global_requests_config.get('download_timeout', 60))

@patch('requests.Session.get')
def test_download_file_from_html_page_success(mock_get, nyfed_connector):
    """測試從 HTML 頁面中找到並下載 Excel 檔案成功。"""
    # 模擬 HTML 頁面內容
    html_content = """
    <html><body>
        <a href="/files/data_file_2023.xlsx">Download Excel Here</a>
        <a href="/files/other_file.pdf">PDF Link</a>
    </body></html>
    """.encode('utf-8')

    # 模擬 Excel 檔案內容
    mock_excel_content = create_mock_excel_bytes({"DataSheet": [["Value"], [100]]})

    # 設定 mock_get 的多個返回值 (第一次是 HTML，第二次是 Excel)
    mock_html_response = MagicMock()
    mock_html_response.status_code = 200
    mock_html_response.headers = {'Content-Type': 'text/html; charset=utf-8'}
    mock_html_response.content = html_content

    mock_excel_response = MagicMock()
    mock_excel_response.status_code = 200
    # 即使 Content-Type 不完全是 Excel，只要包含相關字詞也應能處理
    mock_excel_response.headers = {'Content-Type': 'application/octet-stream'}
    mock_excel_response.content = mock_excel_content.getvalue()

    mock_get.side_effect = [mock_html_response, mock_excel_response]

    page_url = "https://www.example.com/some_page.html"
    file_pattern = "data_file_2023.xlsx" # 要在 HTML 中尋找的檔案模式

    result_bytesio = nyfed_connector._download_file_content(page_url, file_pattern_hint=file_pattern)

    assert result_bytesio is not None
    assert result_bytesio.getvalue() == mock_excel_content.getvalue()
    assert mock_get.call_count == 2
    # 驗證第一次調用 (獲取 HTML)
    mock_get.assert_any_call(page_url, timeout=nyfed_connector.global_requests_config.get('download_timeout', 60))
    # 驗證第二次調用 (下載 Excel) - URL 應是從 HTML 解析出來的絕對 URL
    expected_excel_url = urljoin(page_url, "/files/data_file_2023.xlsx")
    mock_get.assert_any_call(expected_excel_url, timeout=nyfed_connector.global_requests_config.get('download_timeout', 60))


@patch('requests.Session.get')
def test_download_file_http_404_error(mock_get, nyfed_connector):
    """測試下載時遇到 HTTP 404 錯誤。"""
    mock_response = MagicMock()
    mock_response.status_code = 404
    mock_response.raise_for_status.side_effect = requests.exceptions.HTTPError(response=mock_response)
    mock_get.return_value = mock_response

    url = "https://www.example.com/not_found.xlsx"
    result_bytesio = nyfed_connector._download_file_content(url)

    assert result_bytesio is None # 404 通常不重試，直接返回 None
    mock_get.assert_called_once() # 只嘗試一次

@patch('requests.Session.get')
@patch('time.sleep', return_value=None) # Mock time.sleep to speed up tests
def test_download_file_retry_on_500_then_success(mock_sleep, mock_get, nyfed_connector):
    """測試在遇到可重試的 HTTP 500 錯誤後，重試成功。"""
    mock_error_response = MagicMock()
    mock_error_response.status_code = 500
    mock_error_response.raise_for_status.side_effect = requests.exceptions.HTTPError(response=mock_error_response)

    mock_success_excel_content = create_mock_excel_bytes({"S1": [["OK"]]})
    mock_success_response = MagicMock()
    mock_success_response.status_code = 200
    mock_success_response.headers = {'Content-Type': 'application/vnd.ms-excel'}
    mock_success_response.content = mock_success_excel_content.getvalue()

    mock_get.side_effect = [mock_error_response, mock_success_response] # 第一次失敗，第二次成功

    url = "https://www.example.com/temp_error.xlsx"
    result_bytesio = nyfed_connector._download_file_content(url)

    assert result_bytesio is not None
    assert result_bytesio.getvalue() == mock_success_excel_content.getvalue()
    assert mock_get.call_count == 2 # 初始嘗試 + 1 次重試
    mock_sleep.assert_called_once() # 應該有一次退避等待

# --- 測試 get_configured_data ---
@patch.object(NYFedConnector, '_download_file_content') # Mock 下載方法
def test_get_configured_data_success_one_config(mock_download, nyfed_connector, nyfed_config_with_recipes):
    """測試成功處理一個下載設定。"""
    # 準備模擬的 Excel 內容
    recipe_name = "test_dealer_recipe"
    recipe = nyfed_config_with_recipes["parser_recipes"][recipe_name]

    excel_data = [
        ["Some other header"], # 假設 header_row = 2 (1-based), 所以這是第0行 (0-indexed)
        [recipe["date_column"], "Metric A", "Metric B", "Other Col"], # 這是第1行 (標頭)
        [datetime(2023, 1, 1), 10, 20, 99],
        [datetime(2023, 1, 2), 15, 25, 88]
    ]
    mock_excel_bytes = create_mock_excel_bytes({recipe.get("sheet_name","Sheet1"): excel_data})

    # 設定 _download_file_content 的返回值
    # 只 mock 第一個 download_config 的下載，第二個讓它返回 None
    def download_side_effect(url, file_pattern_hint=None):
        if "prideal" in url: # 假設這是第一個 config 的 URL 特徵
            return mock_excel_bytes
        return None # 其他 config 下載失敗

    mock_download.side_effect = download_side_effect

    result_df = nyfed_connector.get_configured_data()

    assert not result_df.empty
    assert len(result_df) == 2
    assert result_df['metric_value'].tolist() == [30.0, 40.0] # 10+20, 15+25
    assert result_df['metric_name'].iloc[0] == "NYFED/TEST_DEALER_METRIC"
    assert result_df['metric_date'].iloc[0] == date(2023, 1, 1)

    # 驗證 _download_file_content 被調用了兩次 (因為有兩個 download_configs)
    assert mock_download.call_count == len(nyfed_config_with_recipes["download_configs"])


@patch.object(NYFedConnector, '_download_file_content', return_value=None) # 所有下載都失敗
def test_get_configured_data_all_downloads_fail(mock_download, nyfed_connector):
    """測試所有檔案下載都失敗的情況。"""
    result_df = nyfed_connector.get_configured_data()
    assert result_df.empty
    assert list(result_df.columns) == nyfed_connector._get_standard_columns()

def test_get_configured_data_no_configs(nyfed_config_minimal):
    """測試設定檔中沒有 download_configs 的情況。"""
    connector_no_configs = NYFedConnector(api_config=nyfed_config_minimal)
    result_df = connector_no_configs.get_configured_data()
    assert result_df.empty

@patch.object(NYFedConnector, '_download_file_content')
def test_get_configured_data_recipe_not_found(mock_download, nyfed_config_with_recipes):
    """測試解析配方未找到的情況。"""
    # 模擬下載成功
    mock_excel_bytes = create_mock_excel_bytes({"Sheet1": [["Date"], [datetime(2023,1,1)]]})
    mock_download.return_value = mock_excel_bytes

    # 建立一個 config，其中 parser_recipe_name 指向一個不存在的 recipe
    bad_config = nyfed_config_with_recipes.copy()
    bad_config["download_configs"] = [
        {
            "name": "bad_recipe_test",
            "url_template": "https://www.example.com/file.xlsx",
            "parser_recipe_name": "NON_EXISTENT_RECIPE", # 這個配方不存在
            "metric_name_override": "NYFED/BAD_RECIPE_METRIC"
        }
    ]
    connector_bad_recipe = NYFedConnector(api_config=bad_config)
    result_df = connector_bad_recipe.get_configured_data()

    assert result_df.empty # 因為配方找不到，應該跳過處理並返回空


@patch.object(NYFedConnector, '_download_file_content')
def test_get_configured_data_parsing_error_date_column_missing(mock_download, nyfed_connector, nyfed_config_with_recipes):
    """測試 Excel 檔案中缺少日期欄位導致的解析錯誤。"""
    recipe_name = "test_dealer_recipe" # 使用這個配方，它期望 "Report Date"
    recipe = nyfed_config_with_recipes["parser_recipes"][recipe_name]

    # 創建一個缺少 "Report Date" 欄位的 Excel
    excel_data_no_date_col = [
        ["Header1", "Metric A", "Metric B"], # 標頭行，但缺少 "Report Date"
        ["Value1", 10, 20]
    ]
    mock_excel_bytes = create_mock_excel_bytes({recipe.get("sheet_name","Sheet1"): excel_data_no_date_col})
    mock_download.return_value = mock_excel_bytes # 讓所有下載都返回這個壞檔案

    result_df = nyfed_connector.get_configured_data()
    assert result_df.empty # 應該因為解析錯誤而返回空 (或只包含其他成功的 config 的數據)
                           # 在這個測試中，因為 mock_download 對所有調用都返回壞檔案，所以最終是空。

# 可以添加更多測試，例如：
# - columns_to_sum 中的欄位不存在於 Excel 中
# - value_column 不存在於 Excel 中
# - 數據單位轉換是否正確
# - CSV 檔案的解析 (如果加入了 CSV 支持)
# - _download_file_content 對於不同 Content-Type 的處理 (例如 text/plain 但實際是 Excel)
# - 速率限制是否按預期工作 (需要更複雜的 time.sleep mock)
# - 年份模板 {YYYY} 是否被正確替換
# - file_pattern_on_page 的不同匹配情況
